"""URRT v8.2 workbook reader. Deterministic, zero LLM tokens.

Oregon is the only state found publishing the raw Unified Rate Review Template
(source-recon.md §2), and that workbook is the load-bearing reason Oregon was
selected (ADR 0001 §1). Worksheet 2 is the native grain of the fact table, so
these 66 plan rows are the one part of the corpus that needs no model at all.

Three properties of the real files shape this module:

**Worksheet 2 is column-oriented.** Plans run ACROSS columns (E, F, G, ... AF) and
fields run DOWN rows. It is a transpose of the table you want. Everything here
reads a field row once and fans it out across plan columns.

**Rows are located by field number, not by row index.** Column B carries the URRT
field number ('1.4', '2.12', '3.11') and column C the label. Hardcoding row 13 for
Plan ID would silently read the wrong row against a v8.3 template or a file with an
inserted row; locating '1.4' cannot. The field numbers are a federal artifact and
are the stable handle here, exactly as `document_role` is in the manifest.

**Computed cells are cached as `#VALUE!`.** Measured in all four Oregon workbooks:
fields 1.12, 1.13, 2.8, 2.15-2.23 and 3.2 hold the literal error string, because
the files ship with macros disabled and formulas uncalculated. These become
`CellError`, never None — see `schema.CellError` for why that distinction is load-
bearing. Field 1.11, the cumulative rate change this project turns on, is a typed
input and is clean in all four.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.extract.schema import CellError, MaybeDecimal

WKSH1 = "Wksh 1 - Market Experience"
WKSH2 = "Wksh 2 - Plan Product Info"
WKSH3 = "Wksh 3 - Rating Areas"

# Excel serial-date epoch (1900 system, with the well-known leap-year offset).
_EXCEL_EPOCH = dt.date(1899, 12, 30)

_ERROR_CELL = re.compile(r"^#(VALUE|REF|DIV/0|N/A|NAME\?|NULL|NUM)!?$", re.I)

# Fields 3.11 and 3.15 are cached as formatted TEXT ('$723.88'), not numbers,
# because the template applies currency formatting inside the formula. Parsing
# them as strings would put '$723.88' into a Decimal column at Phase 4.
_CURRENCY = re.compile(r"^\(?\s*-?\$?\s*[\d,]+(?:\.\d+)?\s*\)?%?$")

# Attributes that must end up numeric regardless of how the cell was cached.
_NUMERIC_ATTRS = frozenset(
    {
        "av_metal_value",
        "cumulative_rate_change_pct",
        "current_premium_pmpm",
        "loss_ratio",
        "adj_factor_av_cs",
        "adj_factor_csr_load",
        "adj_factor_provider_network",
        "adj_factor_non_ehb",
        "adj_factor_admin_expense",
        "adj_factor_taxes_fees",
        "adj_factor_profit_risk",
        "adj_factor_catastrophic",
        "plan_adjusted_index_rate",
        "calib_age",
        "calib_geo",
        "calib_tobacco",
        "calibrated_plan_adjusted_index_rate",
    }
)

# Filing-grain fields that live on Worksheet 2 but describe the whole submission.
# 1.13 is `#VALUE!` in all four Oregon workbooks — the reason CellError exists.
FILING_FIELD_MAP: dict[str, str] = {
    "1.12": "product_rate_increase_pct",
    "1.13": "submission_level_rate_increase_pct",
}

# Column B field number -> the PlanRateExtract attribute it populates.
# Only fields this project models are listed; the workbook has more.
PLAN_FIELD_MAP: dict[str, str] = {
    "1.1": "product_name",
    "1.2": "product_id",
    "1.3": "plan_name",
    "1.4": "plan_id_hios",
    "1.5": "metal",
    "1.6": "av_metal_value",
    "1.7": "plan_category",
    "1.8": "plan_type",
    "1.9": "on_exchange",
    "1.10": "effective_date",
    "1.11": "cumulative_rate_change_pct",
    "2.12": "current_enrollment",
    "2.13": "current_premium_pmpm",
    "2.14": "loss_ratio",
    "3.3": "adj_factor_av_cs",
    "3.4": "adj_factor_csr_load",
    "3.5": "adj_factor_provider_network",
    "3.6": "adj_factor_non_ehb",
    "3.7": "adj_factor_admin_expense",
    "3.8": "adj_factor_taxes_fees",
    "3.9": "adj_factor_profit_risk",
    "3.10": "adj_factor_catastrophic",
    "3.11": "plan_adjusted_index_rate",
    "3.12": "calib_age",
    "3.13": "calib_geo",
    "3.14": "calib_tobacco",
    "3.15": "calibrated_plan_adjusted_index_rate",
}

# The identity block on Worksheet 1: label in column B, value in the next column.
HEADER_LABELS: dict[str, str] = {
    "company legal name": "company_legal_name",
    "hios issuer id": "hios_issuer_id",
    "state": "state",
    "market": "market",
    "effective date of rate change(s)": "effective_date",
}


class WorkbookError(Exception):
    """The workbook is not a URRT, or its structure is unrecognizable.

    Fatal for one document, never for the batch. The caller turns it into a
    `failed` outcome row naming this class.
    """


@dataclass
class WorkbookHeader:
    """Worksheet 1's identity block.

    This is what closes the Phase 1 handoff's open problem: Oregon's SharePoint
    `Title` is a display label ('Regence BCBS'), and `dim_company` needs a legal
    entity name. `company_legal_name` here is that name, read from a fixed cell —
    no model, no confidence score.
    """

    company_legal_name: str | None = None
    hios_issuer_id: str | None = None
    state: str | None = None
    market: str | None = None
    effective_date: dt.date | None = None
    urrt_version: str | None = None
    locators: dict[str, str] = field(default_factory=dict)


@dataclass
class PlanColumn:
    """One plan's values, keyed by PlanRateExtract attribute name.

    `locators` carries the cell address behind each value so provenance can name
    'Wksh 2!E20' rather than 'the workbook'.
    """

    column: str
    values: dict[str, object] = field(default_factory=dict)
    locators: dict[str, str] = field(default_factory=dict)


@dataclass
class UrrtRead:
    header: WorkbookHeader
    plans: list[PlanColumn]
    rating_areas: dict[str, MaybeDecimal]
    missing_fields: list[str]
    # Filing-grain values off Worksheet 2, with their cell addresses. Values here
    # are routinely CellError — see FILING_FIELD_MAP.
    filing_fields: dict[str, object] = field(default_factory=dict)
    filing_locators: dict[str, str] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Coercion
# ---------------------------------------------------------------------------


def coerce_cell(raw: object, cell_ref: str) -> object:
    """Normalize one cached cell value.

    Returns CellError for a spreadsheet error string, Decimal for numerics, str
    for text, None for genuinely empty. Floats become Decimal via `str()` rather
    than `Decimal(float)` — the workbook stores 0.115161378099627 and the binary
    float expansion would put noise digits into a rate change that Phase 3 then
    compares against a PDF.
    """
    if raw is None:
        return None
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        if _ERROR_CELL.match(text):
            return CellError(raw=text, cell=cell_ref)
        return text
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    if isinstance(raw, int | float | Decimal):
        try:
            return Decimal(str(raw))
        except InvalidOperation:  # pragma: no cover - defensive
            return None
    return raw


def parse_formatted_number(text: str) -> Decimal | None:
    """'$1,254.19' -> Decimal('1254.19'). '(12.5)' -> Decimal('-12.5').

    URRT fields 3.11 and 3.15 cache their formula results as formatted text rather
    than numbers. Without this, a Plan Adjusted Index Rate reaches the warehouse as
    the string '$723.88' and every downstream arithmetic check on it silently
    becomes a string comparison.
    """
    candidate = text.strip()
    if not candidate or not _CURRENCY.match(candidate):
        return None
    negative = candidate.startswith("(") and candidate.endswith(")")
    cleaned = candidate.strip("()").replace("$", "").replace(",", "").replace("%", "").strip()
    if not cleaned or cleaned == "-":
        return None
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -value if negative else value


def excel_serial_to_date(value: object) -> dt.date | None:
    """Worksheet 2 field 1.10 arrives as an Excel serial (46388 == 2027-01-01)."""
    if isinstance(value, dt.date):
        return value
    if isinstance(value, Decimal | int | float):
        try:
            return _EXCEL_EPOCH + dt.timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    return None


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------


def read_urrt(path: Path) -> UrrtRead:
    """Read one URRT workbook. Raises WorkbookError if it is not one."""
    try:
        book = load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    except Exception as exc:  # noqa: BLE001 - re-raised as a named, recordable error
        raise WorkbookError(f"{path.name}: cannot open as a workbook: {exc}") from exc

    missing_sheets = [name for name in (WKSH1, WKSH2) if name not in book.sheetnames]
    if missing_sheets:
        raise WorkbookError(
            f"{path.name}: not a URRT — missing worksheet(s) {missing_sheets}. "
            f"Present: {book.sheetnames}"
        )

    header = _read_header(book[WKSH1])
    plans, missing = _read_plans(book[WKSH2])
    filing_fields, filing_locators = _read_filing_fields(book[WKSH2])
    rating_areas = _read_rating_areas(book[WKSH3]) if WKSH3 in book.sheetnames else {}
    return UrrtRead(
        header=header,
        plans=plans,
        rating_areas=rating_areas,
        missing_fields=missing,
        filing_fields=filing_fields,
        filing_locators=filing_locators,
    )


def _read_filing_fields(sheet: Worksheet) -> tuple[dict[str, object], dict[str, str]]:
    """Read the submission-level fields off Worksheet 2's first value column.

    These describe the whole filing rather than a plan, so only the leftmost value
    column carries them. In all four Oregon workbooks field 1.13 (Submission Level
    Rate Increase %) is `#VALUE!` — the filing's own headline number is NOT
    readable from the workbook and has to come from the PDF. Recording it as a
    CellError is what makes that a documented finding rather than a mystery null.
    """
    field_rows = _index_field_rows(sheet)
    values: dict[str, object] = {}
    locators: dict[str, str] = {}
    for field_no, attr in FILING_FIELD_MAP.items():
        row_idx = field_rows.get(field_no)
        if row_idx is None:
            continue
        for col_idx in range(5, min(sheet.max_column, 40) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            ref = f"{sheet.title}!{cell.coordinate}"
            value = coerce_cell(cell.value, ref)
            if value is None:
                continue
            if isinstance(value, str):
                parsed = parse_formatted_number(value)
                value = parsed if parsed is not None else CellError(raw=value[:64], cell=ref)
            values[attr] = value
            locators[attr] = ref
            break
    return values, locators


def _read_header(sheet: Worksheet) -> WorkbookHeader:
    """Scan the top-left block for 'Label:' / value pairs.

    Scanned rather than hardcoded to `C3`/`C4` because the labels sit in column B
    on Worksheet 1 and column C on Worksheet 2 — the same block, shifted one
    column. Matching on the label survives that; a fixed address does not.
    """
    header = WorkbookHeader()
    for row in sheet.iter_rows(min_row=1, max_row=12, max_col=8):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            label = cell.value.strip().rstrip(":").strip().lower()
            if label.startswith("unified rate review"):
                # The cell trails a long accessibility instruction; keep the version.
                header.urrt_version = re.split(r"\s*\[", cell.value.strip(), maxsplit=1)[0]
                header.locators["urrt_version"] = f"{sheet.title}!{cell.coordinate}"
                continue
            attr = HEADER_LABELS.get(label)
            if attr is None or getattr(header, attr, None) is not None:
                continue
            neighbour = sheet.cell(row=cell.row, column=cell.column + 1)
            value = coerce_cell(neighbour.value, f"{sheet.title}!{neighbour.coordinate}")
            if value is None or isinstance(value, CellError):
                continue
            if attr == "effective_date":
                value = excel_serial_to_date(value) or _parse_date(value)
                if value is None:
                    continue
            elif attr == "hios_issuer_id":
                value = _as_issuer_id(value)
            else:
                value = str(value).strip()
            setattr(header, attr, value)
            header.locators[attr] = f"{sheet.title}!{neighbour.coordinate}"
    return header


def _read_plans(sheet: Worksheet) -> tuple[list[PlanColumn], list[str]]:
    """Transpose Worksheet 2 from column-per-plan into row-per-plan.

    The plan columns are whichever columns carry a value on the field-1.4 row
    (Plan ID). Deriving them from the data rather than assuming 'E onward to the
    last used column' matters because the template leaves labelled but empty
    columns behind when a plan is removed.
    """
    field_rows = _index_field_rows(sheet)

    plan_id_row = field_rows.get("1.4")
    if plan_id_row is None:
        raise WorkbookError(
            f"{sheet.title}: no field 1.4 (Plan ID) row — worksheet structure unrecognized"
        )

    plan_columns: list[int] = []
    for cell in sheet[plan_id_row]:
        if cell.column <= 4:  # A-D hold the field number and label, never a plan
            continue
        value = coerce_cell(cell.value, cell.coordinate)
        if isinstance(value, str) and value.strip():
            plan_columns.append(cell.column)

    if not plan_columns:
        raise WorkbookError(f"{sheet.title}: field 1.4 row carries no plan IDs")

    missing = sorted(set(PLAN_FIELD_MAP) - set(field_rows))

    plans: list[PlanColumn] = []
    for col_idx in plan_columns:
        column_letter = sheet.cell(row=plan_id_row, column=col_idx).column_letter
        plan = PlanColumn(column=column_letter)
        for field_no, attr in PLAN_FIELD_MAP.items():
            row_idx = field_rows.get(field_no)
            if row_idx is None:
                continue
            cell = sheet.cell(row=row_idx, column=col_idx)
            ref = f"{sheet.title}!{cell.coordinate}"
            value = coerce_cell(cell.value, ref)
            if value is None:
                continue
            plan.values[attr] = _normalize_plan_field(attr, value)
            plan.locators[attr] = ref
        plans.append(plan)
    return plans, missing


def _index_field_rows(sheet: Worksheet) -> dict[str, int]:
    """Field number (column B) -> row index. The stable addressing scheme."""
    rows: dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, str):
            token = cell.value.strip()
            if re.fullmatch(r"\d+\.\d+", token) and token not in rows:
                rows[token] = cell.row
        elif isinstance(cell.value, int | float) and cell.value is not None:
            token = str(cell.value).strip()
            if re.fullmatch(r"\d+\.\d+", token) and token not in rows:
                rows[token] = cell.row
    return rows


def _normalize_plan_field(attr: str, value: object) -> object:
    if isinstance(value, CellError):
        return value
    if attr == "on_exchange":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("yes", "y", "true", "1")
    if attr == "effective_date":
        return excel_serial_to_date(value) or _parse_date(value)
    if attr == "current_enrollment":
        return int(value) if isinstance(value, Decimal | int | float) else None
    if attr in _NUMERIC_ATTRS:
        if isinstance(value, Decimal):
            return value
        parsed = parse_formatted_number(str(value))
        # A numeric field holding unparseable text is a source defect, not a
        # value. Surfacing it as CellError keeps it distinguishable from absence.
        return parsed if parsed is not None else CellError(raw=str(value)[:64], cell="")
    if attr in ("product_name", "product_id", "plan_name", "plan_id_hios", "metal",
                "plan_category", "plan_type"):
        return str(value).strip()
    return value


def _read_rating_areas(sheet: Worksheet) -> dict[str, MaybeDecimal]:
    areas: dict[str, MaybeDecimal] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        label, factor = row[0], row[1]
        if isinstance(label.value, str) and label.value.strip().lower().startswith("rating area "):
            value = coerce_cell(factor.value, f"{sheet.title}!{factor.coordinate}")
            if value is not None:
                areas[label.value.strip()] = value  # type: ignore[assignment]
    return areas


def _parse_date(value: object) -> dt.date | None:
    if isinstance(value, dt.date):
        return value
    if not isinstance(value, str):
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _as_issuer_id(value: object) -> str:
    """HIOS issuer IDs are 5 digits and must not lose a leading zero.

    openpyxl hands back a number for a numeric cell, so `str(Decimal('77969'))` is
    fine but a hypothetical issuer `01234` would become `1234`. Zero-padding is
    cheap insurance on a field that joins to federal data.
    """
    if isinstance(value, Decimal | int | float):
        return f"{int(value):05d}"
    return str(value).strip()


__all__ = [
    "FILING_FIELD_MAP",
    "HEADER_LABELS",
    "PLAN_FIELD_MAP",
    "WKSH1",
    "WKSH2",
    "WKSH3",
    "PlanColumn",
    "UrrtRead",
    "WorkbookError",
    "WorkbookHeader",
    "coerce_cell",
    "excel_serial_to_date",
    "parse_formatted_number",
    "read_urrt",
]
