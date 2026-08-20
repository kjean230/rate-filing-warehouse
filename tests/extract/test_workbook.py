"""URRT reader: transposed worksheet, field-number addressing, and #VALUE! cells."""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pipeline.extract.schema import CellError
from pipeline.extract.text.workbook import (
    WorkbookError,
    coerce_cell,
    excel_serial_to_date,
    parse_formatted_number,
    read_urrt,
)
from tests.extract.conftest import make_urrt


def test_reads_the_identity_block(urrt_path: Path):
    """This is what closes the handoff's open problem: Oregon's list Title is a
    display label, and dim_company needs the legal name. It comes from a cell."""
    read = read_urrt(urrt_path)
    assert read.header.company_legal_name == "Test Health Plan, Inc."
    assert read.header.hios_issuer_id == "12345"
    assert read.header.state == "OR"
    assert read.header.market == "Individual"
    assert read.header.effective_date == dt.date(2027, 1, 1)


def test_urrt_version_is_trimmed_of_accessibility_boilerplate(urrt_path: Path):
    assert read_urrt(urrt_path).header.urrt_version == "Unified Rate Review v8.2"


def test_worksheet_two_is_transposed_into_plan_rows(urrt_path: Path):
    """Plans run across columns; fields run down rows. The reader inverts that."""
    read = read_urrt(urrt_path)
    assert len(read.plans) == 2
    assert [p.column for p in read.plans] == ["E", "F"]
    assert read.plans[0].values["plan_id_hios"] == "12345OR0010001"
    assert read.plans[1].values["plan_id_hios"] == "12345OR0010002"


def test_locators_name_the_actual_cell(urrt_path: Path):
    """Provenance must say 'Wksh 2!E19', not 'the workbook'."""
    read = read_urrt(urrt_path)
    locator = read.plans[0].locators["cumulative_rate_change_pct"]
    assert locator.startswith("Wksh 2 - Plan Product Info!E")


def test_rate_change_keeps_precision_as_decimal(urrt_path: Path):
    value = read_urrt(urrt_path).plans[0].values["cumulative_rate_change_pct"]
    assert isinstance(value, Decimal)
    assert value == Decimal("0.1152")


def test_value_error_cells_become_cell_error_not_none(urrt_path: Path):
    """Field 1.13 is `#VALUE!` in ALL FOUR live Oregon workbooks. The filing's own
    headline number is not readable from the workbook, and that has to be a
    recorded fact rather than an indistinguishable null."""
    read = read_urrt(urrt_path)
    value = read.filing_fields["submission_level_rate_increase_pct"]
    assert isinstance(value, CellError)
    assert value.raw == "#VALUE!"
    assert "!E" in value.cell


def test_a_real_submission_rate_is_read_as_a_number(tmp_path: Path):
    path = make_urrt(tmp_path / "ok.xlsm", submission_rate=0.1234)
    value = read_urrt(path).filing_fields["submission_level_rate_increase_pct"]
    assert value == Decimal("0.1234")


def test_currency_formatted_cells_are_parsed_to_decimal(urrt_path: Path):
    """Fields 3.11/3.15 cache as TEXT ('$723.88'). Left as strings they would put
    a currency string into a numeric warehouse column."""
    value = read_urrt(urrt_path).plans[0].values["plan_adjusted_index_rate"]
    assert isinstance(value, Decimal)
    assert value == Decimal("1234.56")


def test_rows_are_found_by_field_number_not_row_index(tmp_path: Path):
    """Insert a row above the data; field-number addressing must still work."""
    path = make_urrt(tmp_path / "shifted.xlsm")
    book = load_workbook(path)
    book["Wksh 2 - Plan Product Info"].insert_rows(9)
    book.save(path)

    read = read_urrt(path)
    assert len(read.plans) == 2
    assert read.plans[0].values["plan_id_hios"] == "12345OR0010001"


def test_plan_columns_come_from_the_data_not_a_fixed_range(tmp_path: Path):
    path = make_urrt(
        tmp_path / "three.xlsm",
        plans=[
            {"plan_id": "12345OR0010001", "rate": 0.10},
            {"plan_id": "12345OR0010002", "rate": 0.11},
            {"plan_id": "12345OR0010003", "rate": 0.12},
        ],
    )
    assert len(read_urrt(path).plans) == 3


def test_exchange_flag_becomes_a_bool(urrt_path: Path):
    assert read_urrt(urrt_path).plans[0].values["on_exchange"] is True


def test_effective_date_handles_an_excel_serial(urrt_path: Path):
    assert excel_serial_to_date(Decimal(46388)) == dt.date(2027, 1, 1)


def test_rating_areas_are_read(urrt_path: Path):
    areas = read_urrt(urrt_path).rating_areas
    assert areas["Rating Area 1"] == Decimal("0.952")


def test_a_non_workbook_raises_a_named_error(corrupt_workbook: Path):
    """Becomes a `failed` outcome row naming WorkbookError, never a swallowed crash."""
    with pytest.raises(WorkbookError, match="cannot open as a workbook"):
        read_urrt(corrupt_workbook)


def test_a_workbook_without_urrt_worksheets_is_refused(tmp_path: Path):
    from openpyxl import Workbook

    path = tmp_path / "not-urrt.xlsm"
    book = Workbook()
    book.active.title = "Sheet1"
    book.save(path)
    with pytest.raises(WorkbookError, match="not a URRT"):
        read_urrt(path)


def test_a_worksheet_without_field_1_4_is_refused(tmp_path: Path):
    path = make_urrt(tmp_path / "no-plan-id.xlsm")
    book = load_workbook(path)
    sheet = book["Wksh 2 - Plan Product Info"]
    for row in sheet.iter_rows(min_col=2, max_col=2):
        if row[0].value == "1.4":
            row[0].value = "9.9"
    book.save(path)
    with pytest.raises(WorkbookError, match="no field 1.4"):
        read_urrt(path)


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("$1,254.19", Decimal("1254.19")),
        ("$723.88", Decimal("723.88")),
        ("(12.5)", Decimal("-12.5")),
        ("12.22%", Decimal("12.22")),
        ("1000", Decimal("1000")),
        ("Renewing", None),
        ("", None),
    ],
)
def test_formatted_number_parsing(text, expected):
    assert parse_formatted_number(text) == expected


@pytest.mark.parametrize("error", ["#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NUM!"])
def test_every_spreadsheet_error_form_is_recognized(error):
    assert isinstance(coerce_cell(error, "Wksh 2!E1"), CellError)


def test_floats_become_decimal_via_string_not_binary_expansion():
    """Decimal(0.1) is 0.1000000000000000055511151231257827; Decimal('0.1') is not."""
    assert coerce_cell(0.1, "A1") == Decimal("0.1")
